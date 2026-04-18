"""Telegram Order Bot v2 — Smart ordering với MongoDB Atlas"""
import os, sys, logging, io, functools
from datetime import date, timedelta
from dotenv import load_dotenv
import r2
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.error import BadRequest
from telegram.ext import (Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes, ConversationHandler)
from openpyxl import load_workbook
import db

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)
load_dotenv()

TOKEN = os.environ.get("BOT_TOKEN")
if not TOKEN:
    logger.critical("BOT_TOKEN required"); sys.exit(1)
EXCEL_PATH = os.environ.get("EXCEL_PATH", "DAILY_ORDER_MIN_xlsx.xlsx")
_raw = os.environ.get("ALLOWED_USER_IDS", "")
ALLOWED_USER_IDS: set[int] = {int(x.strip()) for x in _raw.split(",") if x.strip()}

# ─── Excel buffer: tải từ R2 hoặc đọc từ local ──────────────────────────
EXCEL_BUFFER: io.BytesIO | None = None

def _init_excel_buffer() -> None:
    global EXCEL_BUFFER
    if os.environ.get("R2_ENDPOINT") and os.environ.get("R2_ACCESS_KEY"):
        EXCEL_BUFFER = r2.download_excel()
    else:
        logger.info("R2 not configured, using local Excel file: %s", EXCEL_PATH)

def _get_wb(read_only: bool = False, data_only: bool = False):
    """Trả về Workbook từ buffer R2 hoặc file local."""
    if EXCEL_BUFFER is not None:
        EXCEL_BUFFER.seek(0)
        return load_workbook(EXCEL_BUFFER, read_only=read_only, data_only=data_only)
    return load_workbook(EXCEL_PATH, read_only=read_only, data_only=data_only)

class _Src:
    CAT=1; SUB=2; CODE=3; NAME=4; NCC=9; UNIT=20
class _Out:
    DATE_ROW=4; DATE_COL=12; ITEM_START=18
    STT=1; MA_SP=2; TEN_HANG=3; SO_LUONG=7; DVT=9; NGAY_GIAO=10; NCC=15
SHEET_SRC = "Food T01"
SHEET_OUT = "PR NOODLE"

(ENTRY_POINT, EDITING, EDITING_ITEM, ENTERING_EDIT_QTY,
 CHOOSING_CAT, CHOOSING_ITEM, ENTERING_QTY,
 CHOOSING_HISTORY, ENTERING_HISTORY_DATE,
 CONFIRM_ORDER, ENTERING_DATE, ENTERING_TEMPLATE_NAME) = range(12)

def authorized_only(func):
    @functools.wraps(func)
    async def wrapper(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
        uid = update.effective_user.id if update.effective_user else None
        if ALLOWED_USER_IDS and uid not in ALLOWED_USER_IDS:
            m = update.effective_message
            if m: await m.reply_text("⛔ Không có quyền.")
            return
        return await func(update, ctx)
    return wrapper

def fmt_qty(q: float) -> str:
    return str(int(q)) if q == int(q) else str(q)
def fmt_date(d: date) -> str:
    return d.strftime("%d/%m/%Y")
def date_label(d: date) -> str:
    delta = (d - date.today()).days
    suf = {-1:" (hôm qua)", 0:" (hôm nay)", 1:" (ngày mai)"}.get(delta, "")
    return f"{fmt_date(d)}{suf}"

def load_food_items() -> dict:
    if EXCEL_BUFFER is None and not os.path.exists(EXCEL_PATH):
        logger.critical("Excel not found: %s", EXCEL_PATH); sys.exit(1)
    try:
        wb = _get_wb(read_only=True, data_only=True)
    except Exception as e:
        logger.critical("Cannot open Excel: %s", e); sys.exit(1)
    if SHEET_SRC not in wb.sheetnames:
        logger.critical("Sheet '%s' missing", SHEET_SRC); sys.exit(1)
    items = {}
    for row in wb[SHEET_SRC].iter_rows(min_row=3, values_only=True):
        code, name = row[_Src.CODE], row[_Src.NAME]
        unit, cat, sub, ncc = row[_Src.UNIT], row[_Src.CAT], row[_Src.SUB], row[_Src.NCC]
        if code and name and str(name) != "TÊN SẢN PHẨM":
            items[str(code)] = {
                "code": code, "name": str(name),
                "unit": str(unit) if unit else "kg",
                "cat": str(cat) if cat else "Khác",
                "sub": str(sub) if sub else "",
                "ncc": str(ncc) if ncc else ""}
    wb.close(); logger.info("Loaded %d items", len(items)); return items

def get_cats(items: dict) -> dict:
    cats = {}
    for v in items.values(): cats.setdefault(v["cat"], []).append(v)
    return cats

_init_excel_buffer()  # tải Excel từ R2 trước khi load items
ITEMS = load_food_items()
CATS = get_cats(ITEMS)

def build_order_excel(order_items: list, order_date: date | None = None) -> io.BytesIO:
    if order_date is None: order_date = date.today()
    wb = _get_wb()  # fresh copy từ buffer
    if SHEET_OUT not in wb.sheetnames: raise ValueError(f"Sheet '{SHEET_OUT}' not found")
    ws = wb[SHEET_OUT]
    ws.cell(row=_Out.DATE_ROW, column=_Out.DATE_COL, value=order_date)
    for i, item in enumerate(order_items):
        r = _Out.ITEM_START + i
        ws.cell(row=r, column=_Out.STT,       value=i+1)
        ws.cell(row=r, column=_Out.MA_SP,     value=item["code"])
        ws.cell(row=r, column=_Out.TEN_HANG,  value=item["name"])
        ws.cell(row=r, column=_Out.SO_LUONG,  value=item["qty"])
        ws.cell(row=r, column=_Out.DVT,       value=item["unit"])
        ws.cell(row=r, column=_Out.NGAY_GIAO, value=order_date.strftime("%d/%m/%Y"))
        ws.cell(row=r, column=_Out.NCC,       value=item["ncc"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

async def show_edit_screen(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    order = ctx.user_data.get("order", {})
    btns = []
    for code, item in order.items():
        lbl = f"{item['name']}: {fmt_qty(item['qty'])} {item['unit']}"
        if len(lbl) > 38: lbl = lbl[:35] + "..."
        btns.append([InlineKeyboardButton(f"✏️ {lbl}", callback_data=f"ei:{code}")])
    btns.append([InlineKeyboardButton("➕ Thêm mặt hàng", callback_data="add_item")])
    btns.append([InlineKeyboardButton("✅ Xong – Xác nhận", callback_data="done_editing"),
                 InlineKeyboardButton("💾 Lưu mẫu", callback_data="save_tpl_btn")])
    text = f"✏️ *Đơn hàng ({len(order)} món)* — Chạm để sửa:"
    markup = InlineKeyboardMarkup(btns)
    msg = getattr(update, "message", None) or update.callback_query.message
    try:
        await msg.edit_text(text, reply_markup=markup, parse_mode="Markdown")
    except (BadRequest, AttributeError):
        await msg.reply_text(text, reply_markup=markup, parse_mode="Markdown")
    return EDITING

@authorized_only
async def cmd_order(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    ctx.user_data["order"] = {}; ctx.user_data["order_date"] = date.today()
    recent = db.get_recent_dates(1); tmpls = db.list_templates()
    btns = []
    if recent:
        d = date.fromisoformat(recent[0]); n = len(db.get_order(d) or [])
        btns.append([InlineKeyboardButton(f"🔄 Đơn gần nhất ({fmt_date(d)} — {n} món)", callback_data="en:recent")])
    if len(tmpls) == 1:
        btns.append([InlineKeyboardButton(f"📋 Từ mẫu: {tmpls[0]['name']}", callback_data=f"en:tpl:{tmpls[0]['_id']}")])
    elif len(tmpls) > 1:
        btns.append([InlineKeyboardButton("📋 Từ mẫu đã lưu", callback_data="en:tmpls")])
    btns.append([InlineKeyboardButton("📅 Từ đơn ngày khác", callback_data="en:hist")])
    btns.append([InlineKeyboardButton("✏️ Tạo mới hoàn toàn", callback_data="en:new")])
    await update.message.reply_text("🚀 *Bắt đầu đơn hàng từ đâu?*",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return ENTRY_POINT


async def handle_entry(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer(); v = q.data
    if v == "en:new":
        ctx.user_data["order"] = {}
    elif v == "en:recent":
        dates = db.get_recent_dates(1)
        if dates:
            items = db.get_order(date.fromisoformat(dates[0]))
            ctx.user_data["order"] = {str(it["code"]): it for it in (items or [])}
    elif v == "en:hist":
        return await show_history_menu(update, ctx)
    elif v == "en:tmpls":
        return await show_templates_menu(update, ctx)
    elif v.startswith("en:tpl:"):
        items = db.get_template(v.replace("en:tpl:", ""))
        ctx.user_data["order"] = {str(it["code"]): it for it in (items or [])}
    return await show_edit_screen(update, ctx)

async def show_templates_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tmpls = db.list_templates()
    btns = [[InlineKeyboardButton(f"📋 {t['name']}", callback_data=f"en:tpl:{t['_id']}")] for t in tmpls]
    btns.append([InlineKeyboardButton("↩️ Quay lại", callback_data="en:back_main")])
    await update.callback_query.message.edit_text("📋 *Chọn mẫu:*",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return ENTRY_POINT

async def show_history_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    dates = db.get_recent_dates(7)
    btns = []
    for ds in dates:
        d = date.fromisoformat(ds); n = len(db.get_order(d) or [])
        btns.append([InlineKeyboardButton(f"📅 {fmt_date(d)} — {n} món", callback_data=f"hi:{ds}")])
    btns.append([InlineKeyboardButton("🔍 Nhập ngày cụ thể", callback_data="hi:custom")])
    btns.append([InlineKeyboardButton("↩️ Quay lại", callback_data="hi:back")])
    msg = update.callback_query.message if update.callback_query else update.message
    try:
        await msg.edit_text("📅 *Chọn đơn từ lịch sử:*",
            reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    except (BadRequest, AttributeError):
        await msg.reply_text("📅 *Chọn đơn từ lịch sử:*",
            reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return CHOOSING_HISTORY

async def handle_history(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer(); v = q.data
    if v == "hi:back":
        return await cmd_order_refresh(q.message, ctx)
    if v == "hi:custom":
        await q.message.edit_text("🔍 Nhập ngày *DD/MM/YYYY*\nVD: `25/03/2026`", parse_mode="Markdown")
        return ENTERING_HISTORY_DATE
    ds = v.replace("hi:", "")
    items = db.get_order_by_iso(ds)
    if not items:
        await q.answer("❌ Không tìm thấy đơn ngày này", show_alert=True); return CHOOSING_HISTORY
    ctx.user_data["order"] = {str(it["code"]): it for it in items}
    return await show_edit_screen(update, ctx)

async def cmd_order_refresh(msg, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear(); ctx.user_data["order"] = {}; ctx.user_data["order_date"] = date.today()
    recent = db.get_recent_dates(1); tmpls = db.list_templates(); btns = []
    if recent:
        d = date.fromisoformat(recent[0]); n = len(db.get_order(d) or [])
        btns.append([InlineKeyboardButton(f"🔄 Đơn gần nhất ({fmt_date(d)} — {n} món)", callback_data="en:recent")])
    if tmpls: btns.append([InlineKeyboardButton("📋 Từ mẫu đã lưu", callback_data="en:tmpls")])
    btns.append([InlineKeyboardButton("📅 Từ đơn ngày khác", callback_data="en:hist")])
    btns.append([InlineKeyboardButton("✏️ Tạo mới hoàn toàn", callback_data="en:new")])
    await msg.edit_text("🚀 *Bắt đầu đơn hàng từ đâu?*",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return ENTRY_POINT

async def receive_history_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    try:
        d = date(int(raw[6:10]), int(raw[3:5]), int(raw[0:2]))
    except Exception:
        await update.message.reply_text("⚠️ Sai định dạng. Nhập lại *DD/MM/YYYY*:", parse_mode="Markdown")
        return ENTERING_HISTORY_DATE
    items = db.get_order(d)
    if not items:
        await update.message.reply_text(f"❌ Không có đơn ngày {fmt_date(d)}. Nhập ngày khác:")
        return ENTERING_HISTORY_DATE
    ctx.user_data["order"] = {str(it["code"]): it for it in items}
    return await show_edit_screen(update, ctx)

async def edit_item_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    code = q.data.replace("ei:", "")
    item = ctx.user_data["order"].get(code)
    if not item: return await show_edit_screen(update, ctx)
    ctx.user_data["editing_code"] = code
    btns = [[InlineKeyboardButton(str(i), callback_data=f"eq:{i}") for i in range(1, 6)],
            [InlineKeyboardButton(str(i), callback_data=f"eq:{i}") for i in range(6, 11)],
            [InlineKeyboardButton("✏️ Nhập số khác", callback_data="eq:custom"),
             InlineKeyboardButton("🗑️ Xoá món này", callback_data="eq:remove")],
            [InlineKeyboardButton("↩️ Quay lại", callback_data="eq:back")]]
    await q.message.edit_text(
        f"✏️ *{item['name']}* ({item['unit']})\nHiện tại: *{fmt_qty(item['qty'])}*\nChọn số lượng mới:",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return EDITING_ITEM

async def handle_item_edit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer(); v = q.data.replace("eq:", "")
    code = ctx.user_data.get("editing_code")
    if v == "back":
        return await show_edit_screen(update, ctx)
    if v == "remove":
        ctx.user_data["order"].pop(code, None)
        return await show_edit_screen(update, ctx)
    if v == "custom":
        item = ctx.user_data["order"].get(code, {})
        await q.message.edit_text(
            f"✏️ Nhập số lượng cho *{item.get('name','')}* ({item.get('unit','')}):\n(gõ `0` để xoá)",
            parse_mode="Markdown")
        return ENTERING_EDIT_QTY
    try:
        qty = float(v)
        if code and code in ctx.user_data["order"]:
            ctx.user_data["order"][code]["qty"] = qty
    except ValueError:
        pass
    return await show_edit_screen(update, ctx)

async def receive_edit_qty(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().replace(",", ".")
    try:
        qty = float(raw)
    except ValueError:
        await update.message.reply_text("⚠️ Nhập số hợp lệ:"); return ENTERING_EDIT_QTY
    if qty < 0:
        await update.message.reply_text("⚠️ Không thể âm:"); return ENTERING_EDIT_QTY
    code = ctx.user_data.get("editing_code")
    if code and code in ctx.user_data["order"]:
        if qty == 0:
            ctx.user_data["order"].pop(code)
        else:
            ctx.user_data["order"][code]["qty"] = qty
    return await show_edit_screen(update, ctx)


async def show_cats(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    btns = [[InlineKeyboardButton(f"📁 {cat} ({len(CATS[cat])})", callback_data=f"cat:{cat}")] for cat in CATS]
    btns.append([InlineKeyboardButton("↩️ Quay lại đơn", callback_data="cat:back")])
    msg = getattr(update, "message", None) or update.callback_query.message
    try:
        await msg.edit_text("📋 Chọn nhóm hàng:", reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    except (BadRequest, AttributeError):
        await msg.reply_text("📋 Chọn nhóm hàng:", reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return CHOOSING_CAT

async def show_items(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer(); cat = q.data.replace("cat:", "")
    if cat == "back": return await show_edit_screen(update, ctx)
    ctx.user_data["current_cat"] = cat
    items = CATS.get(cat, [])
    btns = []
    for it in items:
        lbl = f"{it['name']} ({it['unit']})"
        if len(lbl) > 40: lbl = lbl[:37] + "..."
        btns.append([InlineKeyboardButton(lbl, callback_data=f"item:{it['code']}")])
    btns.append([InlineKeyboardButton("⬅️ Quay lại", callback_data="cat:back")])
    await q.message.edit_text(f"📁 *{cat}*\nChọn mặt hàng:",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return CHOOSING_ITEM

async def ask_qty(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    code = q.data.replace("item:", "")
    item = ITEMS.get(code)
    if not item:
        await q.answer("Không tìm thấy mặt hàng!", show_alert=True); return CHOOSING_ITEM
    ctx.user_data["current_item"] = item
    existing = ctx.user_data["order"].get(code, {}).get("qty")
    hint = f" (hiện: {fmt_qty(existing)})" if existing is not None else ""
    await q.message.reply_text(
        f"✏️ *{item['name']}* ({item['unit']}){hint}\nNhập số lượng (0 = bỏ):",
        parse_mode="Markdown")
    return ENTERING_QTY

async def receive_qty(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip().replace(",", ".")
    try:
        qty = float(raw)
    except ValueError:
        await update.message.reply_text("⚠️ Nhập số hợp lệ:"); return ENTERING_QTY
    if qty < 0:
        await update.message.reply_text("⚠️ Không thể âm:"); return ENTERING_QTY
    item = ctx.user_data.get("current_item")
    if not item:
        await update.message.reply_text("❌ Lỗi phiên. Gõ /order lại."); return ConversationHandler.END
    code = str(item["code"])
    if qty == 0:
        ctx.user_data["order"].pop(code, None)
    else:
        ctx.user_data["order"][code] = {"code": item["code"], "name": item["name"],
                                         "qty": qty, "unit": item["unit"], "ncc": item["ncc"]}
    return await show_edit_screen(update, ctx)

async def done_editing(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    order = ctx.user_data.get("order", {})
    if not order:
        await q.message.reply_text("🛒 Đơn trống! Thêm mặt hàng trước."); return EDITING
    order_date = ctx.user_data.get("order_date", date.today())
    text, markup = _confirm_markup(order, order_date)
    await q.message.reply_text(text, reply_markup=markup, parse_mode="Markdown")
    return CONFIRM_ORDER

async def save_tpl_prompt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    tmpls = db.list_templates()
    if tmpls:
        btns = [[InlineKeyboardButton(f"📋 Ghi đè: {t['name']}", callback_data=f"tpl_ow:{t['_id']}")] for t in tmpls]
        btns.append([InlineKeyboardButton("✨ Tạo mẫu mới", callback_data="tpl_new")])
        btns.append([InlineKeyboardButton("↩️ Huỷ", callback_data="tpl_cancel")])
        await q.message.reply_text("💾 *Lưu mẫu:*", reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
        return ENTERING_TEMPLATE_NAME
    await q.message.reply_text("💾 Nhập tên mẫu (VD: *Đơn thường*):", parse_mode="Markdown")
    ctx.user_data["tpl_action"] = "new"
    return ENTERING_TEMPLATE_NAME

async def handle_tpl_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer(); v = q.data
    if v == "tpl_cancel": return await show_edit_screen(update, ctx)
    if v == "tpl_new":
        await q.message.reply_text("✨ Nhập tên mẫu mới:")
        ctx.user_data["tpl_action"] = "new"; return ENTERING_TEMPLATE_NAME
    if v.startswith("tpl_ow:"):
        name = v.replace("tpl_ow:", "")
        db.save_template(name, list(ctx.user_data["order"].values()))
        await q.message.reply_text(f"✅ Đã cập nhật mẫu *{name}*!", parse_mode="Markdown")
        return await show_edit_screen(update, ctx)
    return ENTERING_TEMPLATE_NAME

async def receive_tpl_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("⚠️ Tên không được để trống:"); return ENTERING_TEMPLATE_NAME
    db.save_template(name, list(ctx.user_data["order"].values()))
    await update.message.reply_text(f"✅ Đã lưu mẫu *{name}*!", parse_mode="Markdown")
    return await show_edit_screen(update, ctx)

def _confirm_markup(order: dict, order_date: date):
    lines = ["📋 *Xác nhận đơn đặt hàng:*\n"]
    for v in order.values():
        lines.append(f"  • {v['name']}: *{fmt_qty(v['qty'])} {v['unit']}*")
    lines.append(f"\n📅 *Ngày:* {date_label(order_date)}\n_Tổng: {len(order)} mặt hàng_")
    btns = [[InlineKeyboardButton("✅ Tạo file Excel", callback_data="confirm_yes"),
             InlineKeyboardButton("📅 Đổi ngày", callback_data="change_date")],
            [InlineKeyboardButton("✏️ Sửa tiếp", callback_data="back_to_edit"),
             InlineKeyboardButton("❌ Huỷ", callback_data="confirm_no")]]
    return "\n".join(lines), InlineKeyboardMarkup(btns)

async def confirm_yes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer("⏳ Đang tạo file...")
    order = ctx.user_data.get("order", {})
    order_date = ctx.user_data.get("order_date", date.today())
    items_list = list(order.values())
    try:
        buf = build_order_excel(items_list, order_date)
        db.save_order(order_date, items_list)
        filename = f"DonDatHang_{order_date.strftime('%d%m%Y')}.xlsx"
        await q.message.reply_document(document=buf, filename=filename,
            caption=(f"✅ *File đặt hàng ngày {fmt_date(order_date)}*\n"
                     f"📦 {len(items_list)} mặt hàng\n_Gửi cho nhà cung cấp!_"),
            parse_mode="Markdown")
    except Exception:
        logger.exception("Error building order"); await q.message.reply_text("❌ Lỗi tạo file. Thử lại.")
    ctx.user_data.clear(); return ConversationHandler.END

async def confirm_no(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    await q.message.reply_text("❌ Đã huỷ. Gõ /order để bắt đầu lại.")
    ctx.user_data.clear(); return ConversationHandler.END

async def back_to_edit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    return await show_edit_screen(update, ctx)

async def change_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    today = date.today(); yesterday = today - timedelta(days=1); tomorrow = today + timedelta(days=1)
    btns = [
        [InlineKeyboardButton(f"⬅️ {fmt_date(yesterday)} (hôm qua)", callback_data=f"qdate:{yesterday.isoformat()}")],
        [InlineKeyboardButton(f"📅 {fmt_date(today)} (hôm nay)", callback_data=f"qdate:{today.isoformat()}")],
        [InlineKeyboardButton(f"➡️ {fmt_date(tomorrow)} (ngày mai)", callback_data=f"qdate:{tomorrow.isoformat()}")],
        [InlineKeyboardButton("✏️ Nhập ngày khác (DD/MM/YYYY)", callback_data="qdate:custom")],
        [InlineKeyboardButton("↩️ Quay lại", callback_data="qdate:back")]]
    await q.message.edit_text("📅 *Chọn ngày đặt hàng:*",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown")
    return CONFIRM_ORDER

async def quick_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer(); v = q.data.replace("qdate:", "")
    if v == "custom":
        await q.message.edit_text("✏️ Nhập ngày *DD/MM/YYYY*\nVD: `25/12/2025`", parse_mode="Markdown")
        return ENTERING_DATE
    order = ctx.user_data.get("order", {})
    if v == "back":
        order_date = ctx.user_data.get("order_date", date.today())
        text, markup = _confirm_markup(order, order_date)
        await q.message.edit_text(text, reply_markup=markup, parse_mode="Markdown")
        return CONFIRM_ORDER
    chosen = date.fromisoformat(v); ctx.user_data["order_date"] = chosen
    text, markup = _confirm_markup(order, chosen)
    await q.message.edit_text(text, reply_markup=markup, parse_mode="Markdown")
    return CONFIRM_ORDER

async def enter_custom_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    try:
        chosen = date(int(raw[6:10]), int(raw[3:5]), int(raw[0:2]))
    except Exception:
        await update.message.reply_text("⚠️ Sai định dạng. Nhập lại *DD/MM/YYYY*:", parse_mode="Markdown")
        return ENTERING_DATE
    ctx.user_data["order_date"] = chosen
    text, markup = _confirm_markup(ctx.user_data.get("order", {}), chosen)
    await update.message.reply_text(text, reply_markup=markup, parse_mode="Markdown")
    return CONFIRM_ORDER

async def cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text("❌ Đã huỷ. Gõ /order để bắt đầu lại.")
    return ConversationHandler.END

@authorized_only
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()

    # ─── R2 status ───────────────────────────────────────────────
    if EXCEL_BUFFER is not None:
        size_kb = round(EXCEL_BUFFER.getbuffer().nbytes / 1024)
        r2_status = f"✅ Đã kết nối ({size_kb} KB, {len(ITEMS)} mặt hàng)"
    else:
        r2_status = "⚠️ Không dùng R2 (đọc local)"

    # ─── MongoDB status ──────────────────────────────────────────
    try:
        db._get_db().command("ping")
        mongo_status = "✅ Đã kết nối"
    except Exception:
        mongo_status = "❌ Không kết nối được"

    await update.message.reply_text(
        "👋 Xin chào! Tôi là *Order Bot* 🤖\n\n"
        "🔗 *Trạng thái kết nối:*\n"
        f"  • R2 Storage: {r2_status}\n"
        f"  • MongoDB: {mongo_status}\n\n"
        "📋 *Lệnh có sẵn:*\n"
        "/order — Tạo đơn đặt hàng\n"
        "/list — Xem danh sách mặt hàng\n"
        "/tim <từ khoá> — Tìm kiếm\n"
        "/cancel — Huỷ", parse_mode="Markdown")


@authorized_only
async def cmd_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    MAX_LEN = 3800
    lines = [f"📦 *Danh sách ({len(ITEMS)} sản phẩm):*"]; buf = []
    for cat, items in CATS.items():
        block = [f"\n*{cat}* ({len(items)} món):"]
        for it in items[:5]: block.append(f"  • `{it['code']}` {it['name']} ({it['unit']})")
        if len(items) > 5: block.append(f"  _...và {len(items)-5} món khác_")
        if len("\n".join(lines + buf + block)) > MAX_LEN:
            await update.message.reply_text("\n".join(lines + buf), parse_mode="Markdown")
            lines = []; buf = block
        else:
            buf.extend(block)
    if lines or buf:
        await update.message.reply_text("\n".join(lines + buf), parse_mode="Markdown")

@authorized_only
async def cmd_search(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ctx.args:
        await update.message.reply_text("Cú pháp: /tim <từ khoá>\nVD: /tim thịt bò"); return
    kw = " ".join(ctx.args).lower()
    found = [v for v in ITEMS.values() if kw in v["name"].lower()]
    if not found:
        await update.message.reply_text(f"❌ Không tìm thấy '{kw}'"); return
    lines = [f"🔍 *Kết quả '{kw}':*\n"]
    for it in found[:20]: lines.append(f"  `{it['code']}` {it['name']} ({it['unit']}) — {it['ncc']}")
    if len(found) > 20: lines.append(f"\n_...và {len(found)-20} kết quả khác_")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

def main():
    from telegram.ext import ConversationHandler
    app = Application.builder().token(TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler("order", cmd_order)],
        states={
            ENTRY_POINT: [
                CallbackQueryHandler(handle_entry, pattern="^en:"),
            ],
            EDITING: [
                CallbackQueryHandler(edit_item_menu,  pattern="^ei:"),
                CallbackQueryHandler(show_cats,       pattern="^add_item$"),
                CallbackQueryHandler(done_editing,    pattern="^done_editing$"),
                CallbackQueryHandler(save_tpl_prompt, pattern="^save_tpl_btn$"),
            ],
            EDITING_ITEM: [
                CallbackQueryHandler(handle_item_edit, pattern="^eq:"),
            ],
            ENTERING_EDIT_QTY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_edit_qty)
            ],
            CHOOSING_CAT: [
                CallbackQueryHandler(show_items, pattern="^cat:"),
            ],
            CHOOSING_ITEM: [
                CallbackQueryHandler(ask_qty,     pattern="^item:"),
                CallbackQueryHandler(show_cats,   pattern="^cat:back$"),
            ],
            ENTERING_QTY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_qty)
            ],
            CHOOSING_HISTORY: [
                CallbackQueryHandler(handle_history, pattern="^hi:"),
            ],
            ENTERING_HISTORY_DATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_history_date)
            ],
            CONFIRM_ORDER: [
                CallbackQueryHandler(confirm_yes,   pattern="^confirm_yes$"),
                CallbackQueryHandler(confirm_no,    pattern="^confirm_no$"),
                CallbackQueryHandler(back_to_edit,  pattern="^back_to_edit$"),
                CallbackQueryHandler(change_date,   pattern="^change_date$"),
                CallbackQueryHandler(quick_date,    pattern="^qdate:"),
            ],
            ENTERING_DATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_custom_date)
            ],
            ENTERING_TEMPLATE_NAME: [
                CallbackQueryHandler(handle_tpl_action, pattern="^tpl_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_tpl_name),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("list",  cmd_list))
    app.add_handler(CommandHandler("tim",   cmd_search))
    app.add_handler(conv)
    logger.info("Bot đang chạy... (%d items loaded)", len(ITEMS))
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()


