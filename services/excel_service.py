"""Excel loading and order file generation service."""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class ExcelConfig:
    """Excel sheet and column configuration."""
    sheet_src: str = "Food T01"
    sheet_out: str = "PR NOODLE"
    # Source column indices (0-based from Excel)
    col_cat: int = 1
    col_sub: int = 2
    col_code: int = 3
    col_name: int = 4
    col_ncc: int = 9
    col_unit: int = 20
    # Output row/column positions
    out_date_row: int = 4
    out_date_col: int = 12
    out_item_start: int = 18
    out_item_end: int = 40
    out_stt: int = 1
    out_ma_sp: int = 2
    out_ten_hang: int = 3
    out_so_luong: int = 7
    out_dvt: int = 9
    out_ngay_giao: int = 10
    out_ncc: int = 15
    # Non-food output row/column positions
    nonfood_out_date_row: int = 13
    nonfood_out_date_col: int = 2
    nonfood_out_item_start: int = 16
    nonfood_out_item_end: int = 26
    nonfood_out_code_col: int = 2
    nonfood_out_qty_col: int = 7


class ExcelService:
    """Handles Excel loading and order file generation."""

    def __init__(self, buffer: io.BytesIO | None = None, local_path: str | None = None):
        self._buffer = buffer
        self._local_path = local_path
        self._config = ExcelConfig()
        self._items: dict | None = None
        self._categories: dict | None = None

    def load_items(self) -> tuple[dict, dict]:
        """Load and parse food items from Excel. Returns (items, categories)."""
        if self._items is not None:
            return self._items, self._categories

        wb = self._get_workbook(read_only=True, data_only=True)
        if self._config.sheet_src not in wb.sheetnames:
            raise ValueError(f"Sheet '{self._config.sheet_src}' not found")

        items = {}
        for row in wb[self._config.sheet_src].iter_rows(min_row=3, values_only=True):
            code, name = row[self._config.col_code], row[self._config.col_name]
            if code and name and str(name) != "TÊN SẢN PHẨM":
                items[str(code)] = {
                    "code": code,
                    "name": str(name),
                    "unit": str(row[self._config.col_unit]) if row[self._config.col_unit] else "kg",
                    "cat": str(row[self._config.col_cat]) if row[self._config.col_cat] else "Khác",
                    "sub": str(row[self._config.col_sub]) if row[self._config.col_sub] else "",
                    "ncc": str(row[self._config.col_ncc]) if row[self._config.col_ncc] else "",
                }
        wb.close()

        categories: dict = {}
        for v in items.values():
            categories.setdefault(v["cat"], []).append(v)

        self._items = items
        self._categories = categories
        logger.info("Loaded %d items", len(items))
        return items, categories

    def build_order_excel(self, order_items: list, order_date: date) -> io.BytesIO:
        """Generate order Excel file from list of order items."""
        wb = self._get_workbook()
        if self._config.sheet_out not in wb.sheetnames:
            raise ValueError(f"Sheet '{self._config.sheet_out}' not found")

        ws = wb[self._config.sheet_out]
        ws.cell(row=self._config.out_date_row, column=self._config.out_date_col, value=order_date)

        # 1. Xoá các dòng thừa (từ 19 đến 24)
        ws.delete_rows(19, 6)

        item_count = len(order_items)
        template_row = self._config.out_item_start # Dòng 18

        # 2. Insert rows và copy style/công thức từ dòng 18
        if item_count > 1:
            ws.insert_rows(template_row + 1, item_count - 1)
            
            from copy import copy
            from openpyxl.formula.translate import Translator
            
            for i in range(1, item_count):
                r = template_row + i
                for c in range(1, ws.max_column + 1):
                    src_cell = ws.cell(row=template_row, column=c)
                    dst_cell = ws.cell(row=r, column=c)
                    
                    if src_cell.has_style:
                        dst_cell.font = copy(src_cell.font)
                        dst_cell.border = copy(src_cell.border)
                        dst_cell.fill = copy(src_cell.fill)
                        dst_cell.number_format = src_cell.number_format
                        dst_cell.protection = copy(src_cell.protection)
                        dst_cell.alignment = copy(src_cell.alignment)
                    
                    if src_cell.data_type == 'f' and src_cell.value:
                        try:
                            dst_cell.value = Translator(src_cell.value, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
                        except Exception as e:
                            logger.error(f"Formula translation error at {src_cell.coordinate}: {e}")
                            dst_cell.value = src_cell.value

        # 3. Đưa data vào các dòng
        for i, item in enumerate(order_items):
            r = template_row + i
            ws.cell(row=r, column=self._config.out_stt, value=i + 1)
            ws.cell(row=r, column=self._config.out_ma_sp, value=item["code"])
            ws.cell(row=r, column=self._config.out_ten_hang, value=item["name"])
            ws.cell(row=r, column=self._config.out_so_luong, value=item["qty"])
            ws.cell(row=r, column=self._config.out_dvt, value=item["unit"])
            ws.cell(row=r, column=self._config.out_ngay_giao, value=order_date.strftime("%d/%m/%Y"))
            ws.cell(row=r, column=self._config.out_ncc, value=item["ncc"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _get_workbook(self, read_only: bool = False, data_only: bool = False):
        if self._buffer is not None:
            self._buffer.seek(0)
            return load_workbook(self._buffer, read_only=read_only, data_only=data_only)
        return load_workbook(self._local_path, read_only=read_only, data_only=data_only)

    # ─── Non-food methods ───────────────────────────────────────────────────────

    def load_items_nonfood(self) -> tuple[dict, dict]:
        """Load and parse non-food items from CCDC and VTTH sheets.

        Returns (items, categories) where each item includes a `source` field
        so handlers can render [CCDC] / [VTTH] badges on item choices.
        """
        wb = self._get_workbook(read_only=True, data_only=True)

        items: dict = {}
        for sheet_name in ("CCDC", "VTTH"):
            if sheet_name not in wb.sheetnames:
                logger.warning("Non-food sheet '%s' not found in workbook", sheet_name)
                continue
            ws = wb[sheet_name]
            # Assuming row 1 is header, data starts at row 2
            # Expected columns: A=stt, B=code, C=name (or similar - adjust as needed)
            # A more robust approach is to detect headers first
            for row in ws.iter_rows(min_row=2, values_only=True):  # Start at row 2 (row 1 is header)
                # Skip completely empty rows
                if not any(cell is not None for cell in row):
                    continue
                # CCDC sheet columns: A=STT, B=CAT, C=MÃ, D=TÊN, E=ĐVT
                # VTTH sheet expected to have same structure
                cat = row[1] if len(row) > 1 else None  # B = CAT
                code = row[2] if len(row) > 2 else None  # C = MÃ
                name = row[3] if len(row) > 3 else None  # D = TÊN
                unit = row[4] if len(row) > 4 else None  # E = ĐVT
                if code and name and str(code).strip() and str(name).strip():
                    code_str = str(code).strip()
                    items[code_str] = {
                        "code": code_str,
                        "name": str(name).strip(),
                        "unit": str(unit).strip() if unit else "cai",
                        "cat": str(cat).strip() if cat else "Khác",
                        "source": sheet_name,
                    }
        wb.close()

        categories: dict = {}
        for v in items.values():
            categories.setdefault(v["cat"], []).append(v)

        logger.info("Loaded %d non-food items from CCDC+VTTH", len(items))
        return items, categories

    def build_order_excel_nonfood(self, order_items: list, order_date: date) -> io.BytesIO:
        """Generate non-food order Excel.

        Writes product code → column B and quantity → column G.
        All other columns (C, E, H, K) contain VLOOKUP formulas and are preserved.
        Template rows 16-26 are preseeded in the source workbook.
        """
        wb = self._get_workbook()
        sheet_name = self._config.sheet_out
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in non-food workbook")

        ws = wb[sheet_name]

        # Write order date
        ws.cell(row=self._config.nonfood_out_date_row, column=self._config.nonfood_out_date_col, value=order_date)

        # Clear only the cells we are about to write (cols B and G) for template rows
        for r in range(self._config.nonfood_out_item_start, self._config.nonfood_out_item_end + 1):
            ws.cell(row=r, column=self._config.nonfood_out_code_col, value=None)
            ws.cell(row=r, column=self._config.nonfood_out_qty_col, value=None)

        # Validate item count against template capacity
        max_items = self._config.nonfood_out_item_end - self._config.nonfood_out_item_start + 1
        if len(order_items) > max_items:
            logger.warning("Order has %d items but template only supports %d; extras will be dropped", len(order_items), max_items)

        # Write items starting at template row
        for i, item in enumerate(order_items):
            r = self._config.nonfood_out_item_start + i
            if r > self._config.nonfood_out_item_end:
                break  # Don't overflow template rows
            ws.cell(row=r, column=self._config.nonfood_out_code_col, value=item["code"])
            ws.cell(row=r, column=self._config.nonfood_out_qty_col, value=item["qty"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
