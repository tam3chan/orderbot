"""Tests for non-food Excel generation via ExcelService.build_order_excel_nonfood."""

from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from services.excel_service import ExcelService


def _make_nf_workbook() -> Workbook:
    """Create a minimal fake non-food workbook with a PR NOODLE sheet and formula cells."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("PR NOODLE")

    # Row 13: date label + value cell (col B = column 2)
    ws.cell(row=13, column=2, value=None)

    # Rows 16-26: template rows with pre-seeded formula strings in cols C and E
    for r in range(16, 27):
        ws.cell(row=r, column=1, value=f"STT:{r}")
        ws.cell(row=r, column=2, value=None)                          # col B = code (to fill)
        ws.cell(row=r, column=3, value=f"=VLOOKUP(B{r},...)")         # col C = formula (preserve)
        ws.cell(row=r, column=4, value=None)
        ws.cell(row=r, column=5, value=f"=VLOOKUP(B{r},...)")         # col E = formula (preserve)
        ws.cell(row=r, column=6, value=None)
        ws.cell(row=r, column=7, value=None)                          # col G = qty (to fill)
        ws.cell(row=r, column=8, value=None)
        ws.cell(row=r, column=9, value=None)
        ws.cell(row=r, column=10, value=None)
        ws.cell(row=r, column=11, value=None)
        ws.cell(row=r, column=12, value=None)

    return wb


def _excel_service_with_fake_wb() -> ExcelService:
    """Build an ExcelService backed by our fake workbook buffer."""
    wb = _make_nf_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return ExcelService(buffer=buf)


def test_build_order_excel_nonfood_writes_code_and_qty():
    """Code goes to col B (index 2), qty goes to col G (index 7), rows 16+."""
    svc = _excel_service_with_fake_wb()

    items = [
        {"code": "NF01", "name": "Giấy A4", "qty": 3, "unit": "ram", "source": "CCDC"},
        {"code": "NF02", "name": "Bút bi", "qty": 5, "unit": "cây", "source": "VTTH"},
    ]

    result_buf = svc.build_order_excel_nonfood(items, date(2026, 4, 21))

    from openpyxl import load_workbook
    wb_check = load_workbook(result_buf)
    ws = wb_check["PR NOODLE"]

    # Date is written as datetime by openpyxl
    cell_date = ws.cell(row=13, column=2).value
    assert isinstance(cell_date, datetime)
    assert cell_date.date() == date(2026, 4, 21)

    # Row 16 → first item
    assert ws.cell(row=16, column=2).value == "NF01"
    assert ws.cell(row=16, column=7).value == 3

    # Row 17 → second item
    assert ws.cell(row=17, column=2).value == "NF02"
    assert ws.cell(row=17, column=7).value == 5

    wb_check.close()


def test_build_order_excel_nonfood_formula_cells_preserved():
    """Columns C, E (formula cells) are NOT cleared; only col B and G are written."""
    svc = _excel_service_with_fake_wb()

    items = [
        {"code": "NF03", "name": "Xà bông", "qty": 2, "unit": "bịch", "source": "CCDC"},
    ]

    result_buf = svc.build_order_excel_nonfood(items, date(2026, 5, 1))
    from openpyxl import load_workbook
    wb_check = load_workbook(result_buf)
    ws = wb_check["PR NOODLE"]

    # Col C (col 3) = formula string should be preserved for row 16
    c16 = ws.cell(row=16, column=3).value
    assert c16 is not None and "VLOOKUP" in str(c16)

    # Col E (col 5) = formula string should be preserved for row 16
    e16 = ws.cell(row=16, column=5).value
    assert e16 is not None and "VLOOKUP" in str(e16)

    # Col B and G are written with our item data
    assert ws.cell(row=16, column=2).value == "NF03"
    assert ws.cell(row=16, column=7).value == 2

    # Row 17 (unused) should be cleared for col B and G
    assert ws.cell(row=17, column=2).value is None
    assert ws.cell(row=17, column=7).value is None

    # But formula cells in row 17 should still have their formulas
    assert ws.cell(row=17, column=3).value is not None

    wb_check.close()
